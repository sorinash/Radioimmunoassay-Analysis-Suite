# -*- coding: utf-8 -*-
"""
Created on Fri Sep  4 11:16:16 2020

@author: Sam Stewart
"""

import pandas as pd
import numpy as np 
import os
from scipy.optimize import curve_fit
import matplotlib.pyplot as plt
from sys import argv
from statsmodels.formula.api import ols
import statsmodels.api as sm
from openpyxl.drawing.image import Image 
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import shelve

#from PyQt5.QtWidgets import QApplication, Qwidget



# This portion of the program gives you the values for how much tritiated cort
# is necessary, as well as how much antibody is necessary. 

def get_initial_values(*number): 
    if not number: 
        number = 'aaa'
        while not number.isnumeric(): # will repeat until user enters a valid number
            number = input("Please input how many samples you have: ")
        number = int(number)
    else: 
        number = number[0]
    tnb = 9 # there are 3 totals, 3 non-standard bindings, and 3 b0s
    standards = 9 * 3 # there are 9 standards, each in triplicate
    qcs = 15 # there are 15 qcs
    going_forward = tnb + standards + qcs
    total_tubes = (going_forward + (3 * number)) # gets samples in triplicate
    userchoice = 'gleegh' 
    firsttime = True
    while userchoice.lower().strip() not in ['y','n','yes','no']:  # will repeat until user provides an answer.
        if firsttime:
            print('Would you like help calculating values for your experiment?')
        else: 
            print('Please enter a valid option.')
        userchoice = input('Please enter either \'y\' for yes or \'n\' for no:\n')
        firsttime = False
    print('Thank you.\n---------\n\n')
    if userchoice == 'y' or userchoice == 'yes': 
        usercpm = 'gleegh'
        cpmfirst = True
        radiation_vol = ''
        tube_error = total_tubes * 1.15 # accounts for human error by adding 15% extra to volume. 
        tube_volume = tube_error * 100
        while not usercpm.isnumeric() and usercpm != '':  # user must provide valid cpm or just press enter
            if cpmfirst: 
                print('Do you have the cpm for 1 ul of your radiation?')
                print('If so, enter the integer portion of your reading.')
            else: 
                print('Please enter a valid number.')
                print('Remember to only enter the non-decimal portions of your test CPM')
            usercpm = input('Please enter your number, or press Enter if you do not have a test CPM:\n')
            cpmfirst = False
        if usercpm == '': 
            print('That\'s okay! Just remember to input your CPM after getting your reading.')
        else:  # calculates the amount of CORT needed for the experiment
            needed_cpm = tube_volume * 125
            radiation_volume = needed_cpm/int(usercpm)
            radiation_gel = tube_volume - radiation_volume
            radiation_vol += 'To make your radiation solution, you will need approximately '
            radiation_vol += str(radiation_volume) + ' in ' + str(radiation_gel)
            radiation_vol += ' of .1% gel PBS. \n  Remember to record how much you actually used.'

        antibody_volume = tube_error/100
        antibody_gel = tube_volume - antibody_volume
        antibody_str = 'To make your antibody solution, you will need ' 
        antibody_str += str(antibody_volume) + ' of 1:12 antibody stock in '
        antibody_str += str(antibody_gel) + ' of .1% gel PBS.'
        outstr = '----------\nLet\'s get your experiment started!'
        outstr += '\n\n' + radiation_vol + '\n\n' + antibody_str + '\n\n---------'
        print(outstr)

# This will create a template file if so desired. 
# You can manually enter your values into there. 

def make_template_file(filename,number): 
    if not isinstance(number,int) or not str(number).isnumeric() or int(number) < 0: 
        checknumber = 'aaa'
        while not checknumber.isnumeric() or int(checknumber) < 0: 
            print("Please enter a number greater than or equal to 0")
            checknumber = input('\n enter here: ')
        number = int(checknumber)
    else: 
        number = int(number)
    
    if filename == '///invalid///': 
        print("Please enter a valid file name without the extension")
        filename = input('enter here: ')
        filename += '.xls'
    dfframe = gen_template(number)
    dfframe.to_excel(filename,header=False,index=False)

# This will generate a pandas dataframe for make_template_file, which includes
# totals, NSBs, B0s, standards and quality control tubes 

def gen_template(number):
    tnb_labels = ['Labels', 'T','T','T','B0','B0','B0','NSB','NSB','NSB'] 
    qc_labels = ['Q' + str(i) for i in range(1,16)]
    standard_labels = ['S' + str((i//3)+1) for i in range(27)] # returns S1, S2, S3, etc. 
    tube_numbers = [str((i//3) + 1) for i in range(3 * number)]
    # tube_numbers will create triplicate values for checknumber. 
    # so if you have 4 samples, it will create "1,1,1,2,2,2,3,3,3,4,4,4"
    all_labels = tnb_labels + standard_labels + qc_labels + tube_numbers
    cpmlist = ['CPM'] + ([''] * (len(all_labels) -1)) # blank values
    outframe = pd.DataFrame(np.array([all_labels,cpmlist])) # creates a dataframe
    return outframe.transpose() #transposes it for output

# This will reshape the dataframe for ease of reading. 
# Each sample will have its triplicate printed on one line. 

def reshape_dataframe(df):
    uniques = df['Labels'].unique() # gets unique values
    outdict = {}
    maxlen = 0
    for i in uniques: 
        allist = list(df[df['Labels'] == i]['CPM']) #gets CPM values from dataframe
        maxlen = max(maxlen,len(allist)) #gets longest possible length of list
        outdict[str(i)] = allist #adds the list to the output dictionary
    for j in outdict.keys(): 
        outdict[j] += [' '] * (maxlen - len(outdict[j])) # adds spaces in empty cells to equalize the values in the dataframe
    newdf = pd.DataFrame(outdict).transpose() # transposes it for ease of reading.
    return newdf
# And now for the meat of this 

# For our radioimmunoassay protocol, we need to interpolate picograms of 
# corticosterone from counts-per-minute measured by the scintillation counter
# to do so, we create a standard curve from known amounts of cort. 
# This curve is calculated by the following equation: 
# Y=Bottom + (Top-Bottom)/(1+10^((LogIC50-X)*HillSlope))
# where: 
# X is the base-10 log of the corticosterone concentration.
# Top is the maximum value
# Bottom is the minimum value
# IC50 is the concentration giving a value halfway between bottom and top
# and hs is the slope of the curve. 

# log_inhibitor_curve_back is the equation for our standard curve
# log_inhibitor_curve_forward is the equation to interpolate values from CPM 
# based on our standard curve.

def log_inhibitor_curve_back(x,top,bottom,ic,hs): 
    return bottom + ((top-bottom)/(1+10**((np.math.log(ic,10)-x)*hs)))
def log_inhibitor_curve_forward(x,top,bottom,ic,hs):
    #print(top)
    #print(bottom)
    return np.math.log(ic,10) - (np.math.log((top-x)/(x-bottom),10))/hs


# This is our method for interpreting the standard curve
# It takes our known cort mass (pgconcentration) in picograms, and takes
# the base-ten log of these values
# it then goes through a provided dataframe and looks for the CPMs 
# given by the standards. 
# it does so by looking up values
# based on their label (namely s1, s2, s3, s4, s5, s6, s7, s8, and s9), 
# averaging them, and adding them to the "ycpm." 

# it then creates the standard curve. 
# It does so by fitting the x values (pgconcentration)  and yvalues (ycpm)
# to log_inhibitor_curve_back 

# it returns our dataframe, numpy arrays of pgconcentration and ycpm, 
# as well as the optimal values for our fit curve (popt) and the covariance of
# popt (pcov) 

def standard_curve(indf):
    standardlabels = ['S'+ str(i) for i in range(1,10)] #creates standard labels
    pgconcentration = [750,500,250,100,50,25,10,5,2.5] # CORT concentration in correspondence to labels.
    pgconcentration = [np.math.log(i,10) for i in pgconcentration] # the standard curve is on a logarithmic scale
    ycpm = []
    for i in standardlabels:
        serieslist = indf.loc[i] # finds the appropriate values and converts them to a list
        seriesaverage = sum(list(serieslist))/3 # gets the average from the values
        ycpm.append(seriesaverage) 
    popt, pcov = curve_fit(log_inhibitor_curve_back,pgconcentration,ycpm, bounds = ([-np.inf,-np.inf,.1,-100],[np.inf,np.inf,100,100]))
    
    # I place a lower bound of .1 on "ic" to prevent the program
    # from attempting to a value of zero to log_inhibitor_curve_back, 
    # because that would result in the program attempting to take the log of 0.
    
    # Using a number any smaller than .1 for the lower bound of ic functions well on Windows,
    # However, it returns an error on Ubuntu. If the program throws an error at log_inhibitor_curve forward,
    # it is likely because curve_fit's ability to manage small numbers across operating systems appears to be
    # inconsistent. I have confirmed the program's functionality on Windows and Ubuntu.
    
    # I place bounds of -100 and 100 on hs in an attempt to prevent curve_fit
    # from using an arbitrarily large slope in its curve-fitting attempts. 
    # If the absolute value of the standard curve slope is greater than 100, 
    # something has gone dreadfully wrong in the RIA itself. 
    return indf, np.array(pgconcentration), np.array(ycpm), popt, pcov

# after the standard curve is taken, we can use it to interpolate pg values
# from the CPM measured during the RIA. 
# we take the values given to us by standard_curve, 
# then create a list of labels (namely QCs and our tube numbers)
# and isolate labels with those numbers from our dataframe (indf)
# we then convert from CPM using log_inhibitor_curve_forward and the values
# given by popt, and raise 10 to the power of each value to convert from a log
# to a linear scale
# we apply this to all possible values in the triplicate, then take the average
# of those three values and place them in a separate column of the dataframe

def interpolate_curve(indf,popt): 
    [t,b,i,h] = list(popt) # the parameters provided by the curve fitting. Top, Bottom, IC50 and h
    qs =  ['Q' + str(i) for i in range(1,16)] 
    tubelabels = [i for i in indf.index if isinstance(i,int) or i.isnumeric()]
    # I use both isinstance and isnumeric to ensure that strings of numbers
    # and numbers will both be accepted. 
    rels = qs + tubelabels 
    outdf = indf.loc[rels]
    licf = lambda x: 10**(log_inhibitor_curve_forward(x,t,b,i,h)) if not isinstance(x,str) else None # takes 10^x given the existing parameters
    # I use a lambda function here to ensure that I won't accidentally attempt
    # to interpolate a blank cell in the dataframe
    outdf['pg0'] = outdf[0].apply(licf) # creates picogram values for each sample 
    outdf['pg1'] = outdf[1].apply(licf)
    outdf['pg2'] = outdf[2].apply(licf)
    outdf['pgave'] = outdf[['pg0','pg1','pg2']].mean(axis=1) # takes the average
    return outdf 

#loads a file into a dataframe

def load_reflist(*filename):
    if not filename: 
        print('Please input a filename')
        filename = input('\n')
    else: 
        filename = filename[0]
    ref_df = pd.read_excel(filename)
    return ref_df

def load_samples(*filename):
    if not filename: 
        print('Please input a filename')
        filename = input('\n')
    else: 
        filename = filename[0]
    ref_df = pd.read_excel(filename,index_col=0)
    return ref_df

# this will merge your file with the reference file

def merge_with_ref(inoutdf,ref):
    outdf = inoutdf
    for i in ref.columns: 
        outdf.loc[:,i] = ref[i].values # requires merging samples from equally-sized dataframes. Need to trim out standards and the like
    return outdf


# Our lab uses plasma from stressed and non-stressed mice in different amounts
# as stressed mice have higher levels of cort than non-stressed mice, and
# attempting to get both in one volume will throw off our measurements. 

# this will go through each column and divide our interpolated values by the 
# appropriate amount to convert from a mass of picograms into a concentration
# of nanograms per milliliter. 
# It will also divide each of these values by 10 to convert them into ug/dL,
# which is what our field uses to measure CORT concentration. 

def stress_interpret(mergedf): 
    stressdf = mergedf
    stressdf.loc[mergedf['Stress'] == 'S','ng/ml_0'] = mergedf['pg0']/.4 # divides stressed values by .4
    stressdf.loc[mergedf['Stress'] == 'NS','ng/ml_0'] = mergedf['pg0']/1.6 # divides nonstress by 1.6
    stressdf.loc[mergedf['Stress'] == 'S','ng/ml_1'] = mergedf['pg1']/.4
    stressdf.loc[mergedf['Stress'] == 'NS','ng/ml_1'] = mergedf['pg1']/1.6
    stressdf.loc[mergedf['Stress'] == 'S','ng/ml_2'] = mergedf['pg2']/.4
    stressdf.loc[mergedf['Stress'] == 'NS','ng/ml_2'] = mergedf['pg2']/1.6
    stressdf.loc[mergedf['Stress'] == 'S','ng/ml_ave'] = mergedf['pgave']/.4
    stressdf.loc[mergedf['Stress'] == 'NS','ng/ml_ave'] = mergedf['pgave']/1.6
    
    stressdf['ug/dl_0'] = stressdf['ng/ml_0']/10 
    stressdf['ug/dl_1'] = stressdf['ng/ml_1']/10
    stressdf['ug/dl_2'] = stressdf['ng/ml_2']/10
    stressdf['ug_per_dl_ave'] = stressdf['ng/ml_ave']/10
    return stressdf


# this just removes the quality controls from our dataframe 

def begone_qc(indf): 
    qs = ['Q' + str(i) for i in range(1,16)]
    qclessdf = indf[indf.index.isin(qs) == False]
    return qclessdf

# In the event that you want values split up by a specific parameter, splitby
# will do so. That way if you want to only measure, say, male mice, you can split
# by sex 

def splitby(indf,splitlist):
    outlist = [x for x in indf.groupby(splitlist)]
    return outlist


#Overplot takes two inputs, namely, a list of variables to group animals by 
# and a list of dataframes (in case you used splitby to generate a list of graphs)
# takes the form of a list of dfs and a list of parameters to group by.
# dfs take the following format: 
# [the_dfs_label,the_df_itself]



def overplot(dflist,measurelist): 
    plotlist = []
    #abbrevdict converts abbreviations used in the reference file to actual words for the graph
    abbrevdict = {'F': 'Female','M': 'Male', 'KO': 'Knockout', 'WT': 'Wild Type', 'S': 'Stressed', 'NS': 'Non-Stressed'} 
    for i in dflist:  # goes through each dataframe provided
        split_id = i[0] #  gets the initial 
        if split_id in abbrevdict.keys():  
            split_id = abbrevdict[split_id] # creates a list of descriptors 
        groupdf = i[1] 
        outstr = 'Differences in ' + str(split_id) + ' mice' #describes which mice are being plotted
        fig = plt.figure(figsize=(10,10)) #makes a figure
        a = fig.add_subplot(111)
        #print(measurelist)
        #print(groupdf.columns)
        i_ave = groupdf.groupby(measurelist)['ug_per_dl_ave'].mean() # plots the averages
        i_sem = groupdf.groupby(measurelist)['ug_per_dl_ave'].sem() # gets the sem for the error bars
        i_ave.plot.bar(yerr = i_sem)
        fig.suptitle(outstr) # titles it with the description above
        a.set_ylabel('Cort Concentration (ug/dL)') # sets the ylabel to ug/dL
        plotlist.append(fig) # adds that to a list of plots to be interpreted later. 
    return plotlist

#get_that_anova takes the dflist, similar to overplot and the parameters and uses sm.stats to interpret them. 
# It does so by creating a string with the format 'ug_per_dl_ave ~ C(parameter1)*C(parameter2)
#ols is able to use this to create a model from our dataframe
# we can then take this model and run it through sm.stats.anova, which returns 
# a table of results, including stats for interactions between variables. 
# it then outputs these values to a list along with their labels and returns the list

def get_that_anova(dflist,parameters):
    outlist = []
    for i in dflist:
        worklabel = i[0]
        workwith = i[1].dropna() # the anova function sometimes gets finicky with None values
        parameterlist = ['C(%s)'%(j) for j in parameters] # creates a list that can be interpreted by ols
        parameterlist = '*'.join(parameterlist)
        olslist = "ug_per_dl_ave ~ " + parameterlist
        paramtest = ols(olslist,workwith).fit()
        testtable = [worklabel, sm.stats.anova_lm(paramtest)]
        outlist.append(testtable)
    return(outlist)
 
    
 # just isolates the values from Q1-Q15 and analyzes takes their std/mean
 
def quality_control_check(indf): 
     qs = ['Q' + str(i) for i in range(1,16)]
     qcsonly = indf[indf.index.isin(qs) == True]
     return qcsonly['pgave'].std()/qcsonly['pgave'].mean()
     

# arguments 
# --dir: sets the default position for the program to run
# -v: just runs get_initial_values and nothing else
# -m: just runs make_template file and nothing else
# -n: a number value for -v and -m 
# -s: reshapes an input 
# -r: is the reference file for a script
# -i: interpolates a reshaped file from its standard curve
# -t: interprets a reshaped file from a reference file
# -p: runs overplot
# -l: splits by a given parameter or list of parameters. 
# -a: reshapes, interprets, and plots from a single file
# -b: interprets and plots from a reshaped file
# -e: saves your output as a file
# --output: names your file

def RIA_wizard(args): 
    manipstr = '_no_manipulation'
    possible_args = ['-v','-m','--num','-n','-s','-r','-i','-p','-a','-b','c','-t','-l','help','--output','-e','--dir'] # a list of options to recognize
    isthere = {} #True or False values to see if an option is there
    aindex = {} # the index of said option.
    isthere['--num'] = False #sets these to False by default, as searching a list doesn't respond well to "--name=number" formatting
    isthere['help'] = False
    anythere = False
    for acheck in possible_args: 
        isthere[acheck] = acheck in args # Yes if the argument is there, No if not
        if isthere[acheck]: # if yes, it adds the index to the aindex dictionary
            aindex[acheck] = args.index(acheck)
            anythere = True
    for bcheck in range(len(args)):  #goes through and specifically seraches for --num and --output
        if '--num' in args[bcheck]: 
            isthere['--num'] = True
            aindex['--num']  = bcheck
        if '--output' in args[bcheck]: 
            isthere['--output'] = True
            aindex['--output'] = bcheck
        if '--dir' in args[bcheck]: 
            anythere = True
            isthere['--dir'] = True
            aindex['--dir'] = bcheck
    
    # see if there's anything missing 
    if anythere == False: 
        print("No valid inputs found. Please type 'help' to see an explanation of the program")
    # look for 'help' first.
    if isthere['help']: 
        print('''This program accepts the following commands:
              --dir: sets the default directory for this program
              -v : Determines the volume of triated Cort and antibody needed
              -m : Creates a template file for you to paste raw values into 
              --num : Provides a number of samples for -v and -m
              -s : Reshapes an input to display triplicate readings on 1 line
              -r : Refers to a reference file needed to analyze/interpret your data
              -n : Your input file. 
              -i : Interpolates your input file based on a standard curve
              -t : Interprets your reshaped file based on values provided by the reference file
              -p : plots your data 
              -a : Reshapes, interprets, and plots from a single file
              -b : interprets and plots from an already-reshaped file
              -c : reshapes and interprets a file
              -e : saves your output as a file
              --output : names that file
              
              Please bear in mind that if you are entering lists of parameters,
              you will need to connect each parameter with commas between and 
              NO spaces
              
              ie "-p genotype,treatment".  ''')
        return
    
    # before starting the program itself, let the user determine their target directory
    shlf = 'user_dir'
    directory = '/$invalid/$'
    if isthere['--dir']: 
        
        dirdex = aindex['--dir']
        directory = args[dirdex].split('=')[1]
        print("Setting default directory to " + directory)
        try: 
            with shelve.open(shlf,writeback=True) as db:
                db[shlf] = directory
        except: 
            print("Shelve does not have permissions to make a database here")
            return
        
    #then try to set the directory to what the user wants
    #print(directory.split('/'))
    #print(os.getcwd().split('\\'))
    
    if isthere['--dir']:
        os.chdir(directory)
    else: 
        try: 
            with shelve.open(shlf) as db: 
                if shlf in list(db.keys()): 
                    directory = db[shlf]
                    os.chdir(directory)
        except: 
                pass
    
    print('Working in ' + os.getcwd())
    
    # First things first, let's see if the number is there
    number = 'aaa'
    if isthere['--num'] : 
        numdex = aindex['--num']
        number = int(args[numdex].split('=')[1]) # isolates the number from the option
    
    # if '-v' is in the arguments, run get_initial values
    
    if isthere['-v']: 
        get_initial_values(number)
        
    # next up, let's see if -m is in the arguments. 
    # if so, let's see if there's a valid file name up there. 
    # if so, use that as the file name. 
    # otherwise, just plow on ahead with make_template_file
    
    
    
    if isthere['-m']: 
        temfile = '///invalid///'
        temdex = aindex['-m'] 
        if temdex + 1 < len(args) and args[temdex+1] not in possible_args: #ensures that a valid filename is after the option
            temfile = args[temdex+1] + '.xlsx'
            make_template_file(temfile,number)
        else: 
            print("Invalid template filename entered. You can fix that in a bit.") 
        make_template_file(temfile,number)
        return
    
    workingdataframe = 'nope'
    
    # once that's done, let's take a look to see if there's a valid input.     
    
    if isthere['-n']:  #gets the input file
        usefile = '///invalid///'
        usedex = aindex['-n'] 
        if usedex + 1 < len(args) and args[usedex+1] not in possible_args: # ensures a valid filename is provided
            usefile = args[usedex+1] + '.xlsx'
            workingdataframe = load_reflist(usefile)
        else: 
            print("Invalid template filename entered. You can fix that in a bit.") 
            try:
                workingdataframe = load_reflist()
            except: 
                print('Something went wrong. Try again.')
                return
    reffile = '///invalid///'
    
    if isthere['-r']:  # basically the same as above but with the reference file

        refdex = aindex['-r'] 
        if refdex + 1 < len(args) and args[refdex+1] not in possible_args:
            reffile = args[refdex+1] + '.xlsx'
            work_reference = load_reflist(reffile)
            
        else: 
            print("Invalid template filename entered. You can fix that in a bit.") 
            try:
                work_reference = load_reflist()
            except: 
                print('Something went wrong. Try again, or type help for assistance')
                return
        
    # if -s is used gets the reshape
    
        
    if isthere['-s']: 
        #ushapedex = aindex['-s']
        if isthere['-n']: # requires an input to function
            shapedf = workingdataframe
            workingdataframe = reshape_dataframe(shapedf)
            manipstr = '_reshaped'
        else: 
            print('Please enter a valid input file')
            return
    
    my_qc = None
    
    # checks for interpretation
    if isthere['-i']: 
        if isthere['-n']: 
            print('interpreting')
            
            [indf, pgconcentration, ycpm, popt, pcov] = standard_curve(workingdataframe) # we only really care about indf and popt
            workingdataframe = interpolate_curve(indf,popt) # interpolates values from this
            my_qc = quality_control_check(workingdataframe) # also gets the quality control since this is the first time we can do so
            manipstr = "_interpolated" # adds this value to alter the ultimate filename
            # workingdataframe.to_excel('okay_giving_this_a_go.xlsx')
        else: 
            print('Please enter a valid input file')
            return
    
    qlessdataframe = None
    
    if isthere['-t']: 
        #latedex = aindex['-t']
        if isthere['-n'] and isthere['-r']:  # requires input and reference file
            qs =  ['Q' + str(i) for i in range(1,16)]
            tubelabels = [i for i in indf.index if isinstance(i,int) or i.isnumeric()]
            rels = qs + tubelabels
            outdf = workingdataframe.loc[rels] # isolates only QCs and samples
            workingdataframe = merge_with_ref(outdf,work_reference) #merges files together
            workingdataframe = stress_interpret(workingdataframe)  # interprets merged file
            
            qlessdataframe = begone_qc(workingdataframe) # creates a version with qcs removed
            
            manipstr = "_interpreted" 
    
    framelist = None
    splitstr = ""
    
    if isthere['-l']: # if 
        
        splitdex = aindex['-l']
        if type(qlessdataframe) == type(None):  # using type to avoid ambiguous boolean value of dataframe.  
            qlessdataframe = workingdataframe
            splitstr = "_and_split"
            manipstr = "_split"
        if splitdex + 1 < len(args) and args[splitdex + 1] not in possible_args: 
            try: 
                s_parameters = args[splitdex + 1].split(',')
                framelist = splitby(qlessdataframe, s_parameters)
            except: 
                print("Something went wrong in splitting. Try again. Please bear in mind that you cannot use the splitting function while using the '-a' option.")
                print("Please input commands serially, or type help for more options")
                
                return
    plots = []
    anova_df = None
    
    if isthere['-p']: 
        plotdex = aindex['-p']
        if framelist == None and type(qlessdataframe) == type(None):  # if not split
            framelist = [['',begone_qc(workingdataframe)]] # just gets the plot
        elif framelist == None: 
            framelist = [['',qlessdataframe]]
        if plotdex + 1 < len(args) and args[plotdex + 1] not in possible_args: 
            #try: 
            parameters = args[plotdex + 1].split(',') #creates parameters to group by
            print('plotting')
            plots = overplot(framelist,parameters)
            anova_df = get_that_anova(framelist,parameters) # gets the anova
        manipstr = "_plotted" + splitstr
    
    if isthere['-a']: # Does s,i,t,p in quick succession. Cannot use -l with this
        adex = aindex['-a']
        if isthere['-n'] and isthere['-r']:
            manipstr = "_plotted"
            shapedf = workingdataframe
            workingdataframe = reshape_dataframe(shapedf)
            [indf, pgconcentration, ycpm, popt, pcov] = standard_curve(workingdataframe)
            workingdataframe = interpolate_curve(indf,popt)
            my_qc = quality_control_check(workingdataframe)
            qs =  ['Q' + str(i) for i in range(1,16)]
            tubelabels = [i for i in indf.index if isinstance(i,int) or i.isnumeric()]
            rels = qs + tubelabels
            #print(workingdataframe.head())
            outdf = workingdataframe
            workingdataframe = merge_with_ref(outdf,work_reference)
            workingdataframe = stress_interpret(workingdataframe)
            qlessdataframe = begone_qc(workingdataframe)
            framelist = [['',qlessdataframe]]
            if adex + 1 < len(args) and args[adex + 1] not in possible_args:
                parameters = args[adex + 1].split(',')
                plots = overplot(framelist,parameters)
                anova_df = get_that_anova(framelist,parameters)                
                #return [framelist,parameters]
    output_filename = "default_filename.xlsx" # if no value is provided
    
    if isthere['-n']: 
        former = usefile.split('.')[0] 
        output_filename = former + manipstr + '.xlsx' # manipstr, as altered by above, shows the most recent step in the alterations
    
    if isthere['--output']: # if output is here
        outputdex = aindex['--output']
        output_filename = args[outputdex].split('=')[1] # gets the filename after the value
        if '.xls' not in output_filename: 
            output_filename += '.xlsx' # adds an extension just in case the user forgets
    
    if isthere['-e']: # if requested to save
        tempname = 'eeetempfileeee.xlsx' # creates a temporary file to manipulate
        workingdataframe.to_excel(tempname)
        wb = load_workbook(tempname)
        wb_sheet = wb['Sheet1']
        wb_sheet.title = 'Values and Conditions' # renames sheet 1
        if my_qc != None or anova_df != None: # if any stats are there, creates a sheet for them
            ws = wb.create_sheet('Stats')
        if my_qc != None: 
            ws['A1'] = 'Standard Error of Quality Controls = ' + str(my_qc) # adds the qc standard error
        if anova_df != None:  
            for i in anova_df: # goes through, inserts the label, then inserts the relevant anova table
                label = i[0]    
                dftable = i[1]
                ws.append([label])
                for r in dataframe_to_rows(dftable, index = True, header = True):
                    ws.append(r)
        if plots != []:  # if there are any graphs
            wp = wb.create_sheet('Plots') #creates a new sheet
            tplotname = 'templot'
            for p in range(len(plots)):  # goes through all the graphs
                plotname = tplotname + str(p) + '.png' # creates a temporary filename
                figure = plots[p] #  gets the figure data
                figure.savefig(plotname,dpi = 50) # saves that figure as the temp filename
                img = Image(plotname) # makes an image from it
                cellnum = ((p) * 27)+1 # adds it to the plots sheet in the output file
                wp.add_image(img,'A' + str(cellnum)) 
        os.remove(tempname) # gets rid of the temporary file
        wb.save(output_filename) # saves the altered sheet as the output filename
        if plots != []: 
            for rp in range(len(plots)): # then goes through and deletes the images made by the program
                rplotname = tplotname + str(rp) + '.png'
                os.remove(rplotname)
    print('Done! Hopefully you remembered to use -e') # -e  was a relic from when I was attempting to create a GUI from the program. 
    return


RIA_wizard(argv)
